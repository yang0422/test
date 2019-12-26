__author__ = 'xuan'

from openpyxl import load_workbook  # 导入模块
from Common import project_path
import sys
import xlwt;
import xlrd;
from xlutils.copy import copy
import re
import json
from Common.http_request import HttpRequest


class DoExcel:
    def __init__(self, file_path, sheet_name):
        self.file_path = file_path
        self.sheet_name = sheet_name

    def write_data(self, row, actual, result):
        b = load_workbook(self.file_path)
        sheet = b[self.sheet_name]
        sheet.cell(row, 11).value = actual
        sheet.cell(row, 12).value = result
        b.save(self.file_path)

    def write_sql(self, row, check_sql_actual, check_sql_result):
        sql = load_workbook(self.file_path)
        sheet = sql[self.sheet_name]
        sheet.cell(row, 13).value = check_sql_actual
        sheet.cell(row, 14).value = check_sql_result
        sql.save(self.file_path)

    def Write_pass(self, url, method, param, actual, result):
        rexcel = xlrd.open_workbook(self.file_path)
        rows = rexcel.sheets()[self.sheet_name].nrows
        excel = copy(rexcel)
        table = excel.get_sheet(0)
        row = rows
        styleBlueBkg1 = xlwt.easyxf('font: color-index red, bold on')
        table.write(row, 0, url)
        table.write(row, 1, method)
        table.write(row, 2, param)
        table.write(row, 3, actual)
        table.write(row, 4, result, styleBlueBkg1)
        row += 1
        excel.save(self.file_path)

    def Write_fail(self, url, method, param, actual, result):
        rexcel = xlrd.open_workbook(self.file_path)
        rows = rexcel.sheets()[self.sheet_name].nrows
        excel = copy(rexcel)
        table = excel.get_sheet(1)
        row = rows
        styleBlueBkg1 = xlwt.easyxf('font: color-index red, bold on')
        table.write(row, 0, url)
        table.write(row, 1, method)
        table.write(row, 2, param)
        table.write(row, 3, actual)
        table.write(row, 4, result, styleBlueBkg1)
        row += 1
        excel.save(self.file_path)

    def read_file(self):
        a = load_workbook(self.file_path)  # 读取表名
        sheet = a[self.sheet_name]  # 读取表单
        case_list = []
        for i in range(2, sheet.max_row + 1):
            test_data = {}
            test_data['case_id'] = sheet.cell(i, 1).value
            test_data['module'] = sheet.cell(i, 2).value
            test_data['description'] = sheet.cell(i, 3).value
            test_data['method'] = sheet.cell(i, 4).value
            test_data['url'] = sheet.cell(i, 5).value
            test_data['param'] = sheet.cell(i, 6).value
            test_data['expectresult'] = sheet.cell(i, 7).value
            test_data['get_sql'] = sheet.cell(i, 8).value
            test_data['up_sql'] = sheet.cell(i, 9).value
            test_data['check_sql'] = sheet.cell(i, 10).value
            case_list.append(test_data)
        return case_list


if __name__ == '__main__':
    do = DoExcel(project_path.test_data_path, "test_data").read_file()
    for all_data in do:
        print(all_data['description'])







